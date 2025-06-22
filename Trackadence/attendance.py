import Jetson.GPIO as GPIO
import time
import threading
import os
import cv2
import torch
import pickle
import requests
import datetime
import subprocess
import pandas as pd
import numpy as np
from scipy.spatial.distance import cdist
from dotenv import load_dotenv
from retinaface import RetinaFace
from openpyxl import load_workbook
from facenet_pytorch import InceptionResnetV1

# Load environment variables
load_dotenv()

# Define task_running globally
task_running = threading.Event()

# GPIO Pin Definitions
led_pin_1 = 12  # LED Pin
but_pin = 18     # Button Pin

# Initialize Face Recognition Model
recognition_model = InceptionResnetV1(pretrained='vggface2').eval()

def capture_image(directory):
    try:
        os.makedirs(directory, exist_ok=True)
        subprocess.run(["bash", "/home/nvidia/Desktop/Project/capture.sh"], check=True)
        time.sleep(2)
        print("Image captured")
    except Exception as e:
        print(f"Error capturing image: {e}")

def control_relays(blynk_token):
    base_url = "https://blynk.cloud/external/api/"
    try:
        distances = [
            float(requests.get(f"{base_url}get?token={blynk_token}&pin=V{i}").text or 0) 
            for i in range(3)
        ]
        
        relays = [1 if 2 < d < 150 else 0 for d in distances]
        
        for i, value in enumerate(relays, start=3):
            requests.get(f"{base_url}update?token={blynk_token}&pin=V{i}&value={value}")
        
        print("Relay states updated:", dict(zip(["Bulb1", "Bulb2", "Bulb3"], relays)))
    except Exception as e:
        print("Error updating Blynk relays:", e)

def get_face_embedding(face_image):
    face = cv2.resize(face_image, (160, 160))
    face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
    face_tensor = torch.tensor(face).permute(2, 0, 1).float() / 255.0
    face_tensor = face_tensor.unsqueeze(0)
    
    with torch.no_grad():
        return recognition_model(face_tensor).numpy().flatten()

def detect_and_save_faces(image_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    [os.remove(os.path.join(output_dir, f)) for f in os.listdir(output_dir)]
    
    image = cv2.imread(image_path)
    faces = RetinaFace.detect_faces(image_path) if image is not None else {}
    
    for idx, face_data in enumerate(faces.values()):
        x1, y1, x2, y2 = face_data["facial_area"]
        cv2.imwrite(os.path.join(output_dir, f"face_{idx}.jpg"), image[y1:y2, x1:x2])

def get_normalized_embedding(embedding):
    embedding = embedding.flatten()
    return embedding / np.linalg.norm(embedding)

def find_best_match(face_embedding, known_embeddings, threshold=0.4):
    normalized_embedding = get_normalized_embedding(face_embedding)
    distances = cdist([normalized_embedding], known_embeddings, metric='cosine')
    best_match_idx = np.argmin(distances)
    return best_match_idx if (1 - distances[0][best_match_idx]) > threshold else -1

def identify_faces_in_directory(face_dir, embeddings_pk_path, template_path, output_excel):
    with open(embeddings_pk_path, 'rb') as f:
        known_faces = pickle.load(f)
    
    name_list, embedding_list = [], []
    for name, known_embeddings in known_faces.items():
        if isinstance(known_embeddings, list):
            for emb in known_embeddings:
                name_list.append(name)
                embedding_list.append(emb.flatten())
        else:
            name_list.append(name)
            embedding_list.append(known_embeddings.flatten())
    
    known_embeddings = np.vstack(embedding_list)
    current_time = datetime.datetime.now()
    date_str = current_time.strftime("%d-%m-%Y")
    time_str_header = current_time.strftime("%I:%M %p")
    
    workbook = load_workbook(template_path)
    worksheet = workbook.active
    total_present = 0
    
    for face_img_name in os.listdir(face_dir):
        face_path = os.path.join(face_dir, face_img_name)
        face_image = cv2.imread(face_path)
        if face_image is None:
            print(f"Could not read image: {face_path}")
            continue
        
        face_embedding = get_face_embedding(face_image)
        idx = find_best_match(face_embedding, known_embeddings, threshold=0.6)
        identity, usn = ("Unknown", "Unknown") if idx == -1 else name_list[idx].split('_')
        
        if usn != "Unknown":
            for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row):
                usn_cell, attendance_cell = row[2], row[3]
                if usn_cell.value == usn:
                    attendance_cell.value = "P"
                    total_present += 1
                    break
    
    total_students = 10
    total_absent = total_students - total_present
    worksheet.cell(row=4, column=2, value=f"Date: {date_str} Time: {time_str_header}")
    worksheet.cell(row=20, column=2, value=f"Total Students Present: {total_present}")
    worksheet.cell(row=21, column=2, value=f"Total Students Absent: {total_absent}")
    
    workbook.save(output_excel)
    print(f"Attendance updated and saved to {output_excel}")

def send_xlsx_to_telegram(file_path):
    bot_token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("CHAT_ID")
    
    if not bot_token or not chat_id:
        print("Error: Missing Telegram credentials.")
        return
    
    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    caption = f"Attendance report generated at: {current_time}"
    
    try:
        with open(file_path, 'rb') as file:
            response = requests.post(url, data={'chat_id': chat_id, 'caption': caption}, files={'document': file})
            print("File sent successfully to Telegram.")
    except Exception as e:
        print(f"Error sending file to Telegram: {e}")

def main():
    GPIO.setmode(GPIO.BOARD)
    GPIO.setup(led_pin_1, GPIO.OUT)
    GPIO.setup(but_pin, GPIO.IN, pull_up_down=GPIO.PUD_UP)
    GPIO.add_event_detect(but_pin, GPIO.FALLING, callback=button_pressed, bouncetime=300)
    print("Waiting for button press...")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        GPIO.cleanup()
        print("GPIO cleaned up.")

if __name__ == '__main__':
    main()
